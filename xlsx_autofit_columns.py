#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Excel自动调整行高列宽工具

这个脚本用于自动调整Excel文件的行高和列宽，使其内容显示美观清晰。
可以处理中英文混合文本、多行文本，并考虑字体大小等因素。
"""

import os
import sys
import argparse
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import styles
import re


# 字符宽度常量
CHAR_WIDTH = {
    'ascii': 0.7,  # ASCII字符（英文、数字、符号）
    'cjk': 2.0,    # 中日韩字符
    'emoji': 2.0,  # Emoji表情符号
}

# 默认配置
DEFAULT_CONFIG = {
    'width_factor': 1.3,      # 列宽调整系数 (增加以显示更充分)
    'height_factor': 1.3,     # 行高调整系数 (增加以显示更充分)
    'min_width': 8,           # 最小列宽
    'max_width': 120,         # 最大列宽 (增加以显示更充分)
    'min_height': 20,         # 最小行高
    'max_height': 409,        # 最大行高（Excel限制）
    'default_font_size': 12,  # 默认字体大小
    'enable_size_limits': True,    # 是否启用单元格尺寸限制
    'enable_font_autofit': False,   # 是否启用字体大小自适应
    'enable_cell_wrap': True,      # 是否启用单元格自动换行
    'enable_cell_alignment': True,  # 是否启用单元格对齐方式
    'horizontal_alignment': 'center',  # 水平对齐方式：'left', 'center', 'right'
    'vertical_alignment': 'center',    # 垂直对齐方式：'top', 'center', 'bottom'
}

# 默认要处理的Excel文件列表
# 可以直接在此处添加文件路径，程序将自动处理这些文件
DEFAULT_FILES = [
]


def sanitize_filename(filename):
    """处理文件名中的特殊字符"""
    # 返回一个可以安全使用的文件名
    return filename.replace('"', '\"')  # 转义双引号


def get_char_type(char):
    """判断字符类型"""
    code = ord(char)
    if code <= 127:  # ASCII字符
        return 'ascii'
    elif code > 0x1F000:  # Emoji范围
        return 'emoji'
    else:  # 假设其他都是中日韩字符
        return 'cjk'


def calculate_text_width(text, font_size=11):
    """计算文本宽度"""
    if not text:
        return 0
    
    # 将文本转换为字符串
    text = str(text)
    
    # 计算每个字符的宽度并求和
    width = 0
    for char in text:
        char_type = get_char_type(char)
        width += CHAR_WIDTH.get(char_type, CHAR_WIDTH['ascii'])
    
    # 根据字体大小调整宽度
    font_factor = font_size / DEFAULT_CONFIG['default_font_size']
    return width * font_factor


def calculate_text_height(text, column_width, font_size=11):
    """计算文本高度
    
    参数:
        text: 要计算高度的文本
        column_width: 列宽，用于计算自动换行
        font_size: 字体大小
        
    返回:
        考虑了硬换行符和自动换行后的文本高度
    """
    if not text:
        return 1
    
    # 将文本转换为字符串
    text = str(text)
    
    # 处理文本中的硬换行
    lines = text.split('\n')
    
    # 计算每行需要的高度（同时考虑硬换行和自动换行）
    total_lines = 0
    for line in lines:
        if not line:  # 空行也算一行
            total_lines += 1
            continue
            
        # 计算这行文本的宽度
        line_width = calculate_text_width(line, font_size)
        
        # 计算这行文本需要的行数（考虑一定的边距）
        if column_width > 0:
            # 减去一些边距，使文本有更好的显示效果
            effective_width = column_width * 0.9
            line_count = (line_width / effective_width)
            # 向上取整
            line_count = int(line_count) + (1 if line_count > int(line_count) else 0)
            total_lines += max(1, line_count)
        else:
            total_lines += 1
    
    # 根据字体大小调整高度
    font_factor = font_size / DEFAULT_CONFIG['default_font_size']
    return total_lines * font_factor + 1


def calculate_space_utilization(text_size, container_size):
    """计算空间利用率"""
    if container_size <= 0:
        return 1.0
    return text_size / container_size

def adjust_font_size(original_size, utilization, min_factor=0.8, max_factor=1.5):
    """根据空间利用率调整字体大小"""
    if utilization >= 0.5:  # 空间利用率合理，不调整
        return original_size
    
    # 根据利用率计算调整系数
    adjustment_factor = max(min_factor, min(max_factor, 1 / utilization))
    return original_size * adjustment_factor


def autofit_columns(worksheet, config=None):
    """自动调整列宽"""
    if config is None:
        config = DEFAULT_CONFIG
    
    # 获取工作表的数据范围
    data_rows = list(worksheet.rows)
    if not data_rows:
        return
    
    # 遍历所有列
    for col_idx in range(1, worksheet.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_width = 0
        cells_in_column = []
        
        # 遍历该列的所有单元格
        for row_idx in range(1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                # 获取单元格字体大小，如果没有则使用默认值
                font_size = getattr(getattr(cell, 'font', None), 'size', None) or config['default_font_size']
                
                # 计算单元格内容宽度
                text_width = calculate_text_width(cell.value, font_size)
                max_width = max(max_width, text_width)
                
                cells_in_column.append((cell, text_width, font_size))
                
                # 根据配置设置单元格对齐方式和自动换行
                if config['enable_cell_alignment']:
                    cell.alignment = openpyxl.styles.Alignment(
                        horizontal=config['horizontal_alignment'],
                        vertical=config['vertical_alignment'],
                        wrap_text=config['enable_cell_wrap']
                    )
        
        # 应用列宽调整系数
        adjusted_width = max_width * config['width_factor']
        
        # 根据配置决定是否限制列宽
        if config['enable_size_limits']:
            final_width = max(config['min_width'], min(config['max_width'], adjusted_width))
        else:
            final_width = adjusted_width
        
        # 设置列宽
        worksheet.column_dimensions[col_letter].width = final_width
        
        # 根据配置决定是否优化字体大小
        if config['enable_font_autofit']:
            for cell, text_width, original_font_size in cells_in_column:
                utilization = calculate_space_utilization(text_width, final_width)
                new_font_size = adjust_font_size(original_font_size, utilization)
                
                if abs(new_font_size - original_font_size) > 0.1:  # 只在字体大小变化明显时才修改
                    cell.font = openpyxl.styles.Font(
                        size=new_font_size,
                        name=getattr(getattr(cell, 'font', None), 'name', None) or 'Arial'
                    )

def autofit_rows(worksheet, config=None):
    """自动调整行高"""
    if config is None:
        config = DEFAULT_CONFIG
    
    # 遍历所有行
    for row_idx in range(1, worksheet.max_row + 1):
        max_height = 0
        cells_in_row = []
        
        # 遍历该行的所有单元格
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                # 获取单元格字体大小，如果没有则使用默认值
                font_size = getattr(getattr(cell, 'font', None), 'size', None) or config['default_font_size']
                
                # 获取列宽
                col_letter = get_column_letter(col_idx)
                col_width = worksheet.column_dimensions[col_letter].width
                
                # 计算单元格内容高度
                text_height = calculate_text_height(cell.value, col_width, font_size)
                max_height = max(max_height, text_height)
                cells_in_row.append((cell, text_height, font_size, col_width))
        
        # 应用行高调整系数
        adjusted_height = max_height * config['height_factor'] * 15  # 转换为Excel行高单位
        
        # 根据配置决定是否限制行高
        if config['enable_size_limits']:
            final_height = max(config['min_height'], min(config['max_height'], adjusted_height))
        else:
            final_height = adjusted_height
        
        # 设置行高
        worksheet.row_dimensions[row_idx].height = final_height
        
        # 根据配置决定是否优化字体大小
        if config['enable_font_autofit']:
            for cell, text_height, original_font_size, col_width in cells_in_row:
                # 计算空间利用率
                utilization = calculate_space_utilization(text_height * 15, final_height)
                new_font_size = adjust_font_size(original_font_size, utilization)
                
                if abs(new_font_size - original_font_size) > 0.1:  # 只在字体大小变化明显时才修改
                    # 检查新字体大小是否会导致文本过宽
                    text_width = calculate_text_width(cell.value, new_font_size)
                    width_utilization = calculate_space_utilization(text_width, col_width)
                    
                    if width_utilization <= 0.9:  # 确保不会因为增大字体而导致文本过宽
                        cell.font = openpyxl.styles.Font(
                            size=new_font_size,
                            name=getattr(getattr(cell, 'font', None), 'name', None) or 'Arial'
                        )


def process_excel_file(file_path, output_path=None, config=None):
    """处理Excel文件，自动调整行高列宽"""
    # 处理文件名中的特殊字符
    file_path = sanitize_filename(file_path)
    
    if not os.path.exists(file_path):
        print(f"错误：文件 '{file_path}' 不存在")
        return False
    
    # 如果没有指定输出路径，则在原文件名后添加'_beautifuler'后缀
    if output_path is None:
        # 分离文件名和扩展名
        file_name, file_ext = os.path.splitext(file_path)
        output_path = f"{file_name}_beautifuler{file_ext}"
    else:
        output_path = sanitize_filename(output_path)
    
    # 使用默认配置
    if config is None:
        config = DEFAULT_CONFIG
    
    try:
        # 尝试检查文件是否可访问
        try:
            with open(file_path, 'rb') as _:
                pass
        except PermissionError:
            print(f"错误：无法访问文件 '{file_path}'")
            print("可能的原因：")
            print("1. 文件正在被其他程序（如Excel）使用")
            print("2. 没有足够的文件访问权限")
            print("\n建议：")
            print("1. 关闭可能正在使用此文件的其他程序")
            print("2. 确保您有足够的文件访问权限")
            print("3. 尝试将文件复制到其他位置后再处理")
            return False
        
        # 加载工作簿
        print(f"正在处理文件: {file_path}")
        workbook = load_workbook(file_path, read_only=False)
        
        # 处理每个工作表
        for sheet_name in workbook.sheetnames:
            print(f"  调整工作表: {sheet_name}")
            worksheet = workbook[sheet_name]
            
            # 先调整列宽，再调整行高
            autofit_columns(worksheet, config)
            autofit_rows(worksheet, config)
        
        # 尝试保存工作簿
        try:
            workbook.save(output_path)
            print(f"文件已保存: {output_path}")
            return True
        except PermissionError:
            print(f"错误：无法保存文件 '{output_path}'")
            print("可能的原因：")
            print("1. 文件正在被其他程序使用")
            print("2. 没有足够的写入权限")
            print("\n建议：")
            print("1. 关闭可能正在使用此文件的其他程序")
            print("2. 确保您有足够的文件写入权限")
            print("3. 尝试保存到其他位置（使用 -o 参数指定输出路径）")
            return False
    
    except Exception as e:
        print(f"处理文件时出错: {str(e)}")
        print("\n如果问题持续存在，请尝试：")
        print("1. 确保文件格式正确且未损坏")
        print("2. 使用另一个Excel文件测试")
        print("3. 检查文件名是否包含特殊字符")
        return False


def process_default_files(config=None):
    """处理默认文件列表中的所有Excel文件"""
    if not DEFAULT_FILES:
        print("警告：未在代码中设置默认文件路径，请在DEFAULT_FILES列表中添加文件路径或使用命令行参数指定文件")
        return False
    
    success = True
    for file_path in DEFAULT_FILES:
        file_success = process_excel_file(file_path, None, config)
        if not file_success:
            success = False
    
    return success


def main():
    """主函数"""
    # 创建命令行参数解析器
    parser = argparse.ArgumentParser(description='自动调整Excel文件的行高和列宽')
    parser.add_argument('file', nargs='?', help='Excel文件路径（可选，如不提供则使用代码中DEFAULT_FILES列表）')
    parser.add_argument('-o', '--output', help='输出文件路径（默认覆盖原文件）')
    parser.add_argument('--width-factor', type=float, default=DEFAULT_CONFIG['width_factor'], help='列宽调整系数')
    parser.add_argument('--height-factor', type=float, default=DEFAULT_CONFIG['height_factor'], help='行高调整系数')
    parser.add_argument('--min-width', type=float, default=DEFAULT_CONFIG['min_width'], help='最小列宽')
    parser.add_argument('--max-width', type=float, default=DEFAULT_CONFIG['max_width'], help='最大列宽')
    parser.add_argument('--min-height', type=float, default=DEFAULT_CONFIG['min_height'], help='最小行高')
    parser.add_argument('--max-height', type=float, default=DEFAULT_CONFIG['max_height'], help='最大行高')
    parser.add_argument('--font-size', type=float, default=DEFAULT_CONFIG['default_font_size'], help='默认字体大小')
    parser.add_argument('--use-defaults', action='store_true', help='使用代码中DEFAULT_FILES列表中的文件路径')
    
    # 添加功能开关选项
    parser.add_argument('--disable-size-limits', action='store_false', dest='enable_size_limits',
                        help='禁用单元格尺寸限制')
    parser.add_argument('--disable-font-autofit', action='store_false', dest='enable_font_autofit',
                        help='禁用字体大小自适应')
    parser.add_argument('--disable-cell-wrap', action='store_false', dest='enable_cell_wrap',
                        help='禁用单元格自动换行')
    parser.add_argument('--disable-cell-alignment', action='store_false', dest='enable_cell_alignment',
                        help='禁用单元格对齐方式')
    parser.add_argument('--horizontal-alignment', choices=['left', 'center', 'right'],
                        default=DEFAULT_CONFIG['horizontal_alignment'], help='水平对齐方式')
    parser.add_argument('--vertical-alignment', choices=['top', 'center', 'bottom'],
                        default=DEFAULT_CONFIG['vertical_alignment'], help='垂直对齐方式')
    
    # 设置功能开关的默认值
    parser.set_defaults(
        enable_size_limits=DEFAULT_CONFIG['enable_size_limits'],
        enable_font_autofit=DEFAULT_CONFIG['enable_font_autofit'],
        enable_cell_wrap=DEFAULT_CONFIG['enable_cell_wrap'],
        enable_cell_alignment=DEFAULT_CONFIG['enable_cell_alignment']
    )
    
    # 解析命令行参数
    args = parser.parse_args()
    
    # 创建配置
    config = {
        'width_factor': args.width_factor,
        'height_factor': args.height_factor,
        'min_width': args.min_width,
        'max_width': args.max_width,
        'min_height': args.min_height,
        'max_height': args.max_height,
        'default_font_size': args.font_size,
        'enable_size_limits': args.enable_size_limits,
        'enable_font_autofit': args.enable_font_autofit,
        'enable_cell_wrap': args.enable_cell_wrap,
        'enable_cell_alignment': args.enable_cell_alignment,
        'horizontal_alignment': args.horizontal_alignment,
        'vertical_alignment': args.vertical_alignment,
    }
    
    # 处理Excel文件
    success = False
    if args.file:
        # 如果命令行提供了文件路径，则处理该文件
        success = process_excel_file(args.file, args.output, config)
    else:
        # 否则处理默认文件列表
        success = process_default_files(config)
    
    # 返回退出码
    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()